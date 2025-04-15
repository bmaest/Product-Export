import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from datetime import datetime

# API Configuration
base_url = "https://185.koronacloud.com/web/api/v3/accounts"
koronaAccountId = "ae0b738f-1db9-47d4-a4d4-868b37b31d4e"
username = "API"
password = "Combase"
page = 1

# Create a new Excel workbook
workbook = Workbook()
sheet = workbook.active
sheet.title = "Products"

# Define header row
headers = [
    "Product #", "Product Name", "Commodity Group #", "Commodity Group Name",
    "Price Group #", "Price Group Name", "Price Group Value", "Price Validity",
    "Organizational Unit #", "Organizational Unit Name", "Organizational Unit Value",
    "Container #", "Container Name", "Container Price", "Container Price Group #",
    "Default Container", "Container Validity", "containerCapacity",
    "Supplier Number", "Supplier Name", "Order Code", "Supplier Item Price",
    "Container Size", "Product Code", "Container Amount", "Tag",
    "Assortment #", "Assortment Name", "Sector #", "Sector Name", 
    "Alternate Sector #", "Alternate Sector Name", 
    "Item Sequence #", "Item Sequence Name",
    "Deactivated", "Discountable", "Track Inventory"
]
sheet.append(headers)

# Define the fill colors
light_grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Initialize variables for alternating colors
current_fill = light_grey_fill  # Start with light grey
row_index = 2  # Start from the second row, as the first row is the header

# Helper function to fetch data from the API
def fetch_data_from_api(koronaAccountId, page):
    url = f"{base_url}/{koronaAccountId}/products?page={page}"
    print(f"Fetching data from URL: {url}")
    response = requests.get(url, auth=(username, password))
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Failed to fetch data: {response.status_code}")
        return None

# Helper function to process and append product rows
def process_product(product, row_index, current_fill):
    product_name = product.get('name', '')
    product_number = product.get('number', '')
    product_commodity = product.get('commodityGroup', {})
    product_commodity_number = product_commodity.get('number', '')
    product_commodity_name = product_commodity.get('name', '')
    product_container_capacity = product.get('containerCapacity', '')
    product_assortment = product.get('assortment', {})
    product_assortment_number = product_assortment.get('number', '')
    product_assortment_name = product_assortment.get('name', '')
    product_sector = product.get('sector', {})
    product_sector_number = product_sector.get('number', '')
    product_sector_name = product_sector.get('name', '')
    product_altSector = product.get('alternativeSector', {})
    product_altSector_number = product_altSector.get('number', '')
    product_altSector_name = product_altSector.get('name', '')
    product_itemSequence = product.get('itemSequence', {})
    product_itemSequence_number = product_itemSequence.get('number', '')
    product_itemSequence_name = product_itemSequence.get('name', '')
    product_active = not product.get('active', False)
    product_discountable = product.get('discountable', False)
    product_track_inventory = product.get('trackInventory', False)

    # Append the initial row with product basic info
    sheet.append([product_number, product_name, product_commodity_number, product_commodity_name])
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=row_index, column=col)
        cell.fill = current_fill
        cell.alignment = Alignment(horizontal="left")
    row_index += 1

    # Process rows for prices, containers, suppliers, and codes
    product_prices = product.get('prices', [])
    container_prices = [
        {
            "container_name": container.get('product', {}).get('name', ''),
            "container_number": container.get('product', {}).get('number', ''),
            "container_price_value": price.get('value', ''),
            "container_price_group_number": price.get('priceGroup', {}).get('number', ''),
            "default_container": container.get('defaultContainer', False),
            "container_valid": datetime.fromisoformat(price.get('validFrom', '')).strftime('%m/%d/%Y') if price.get('validFrom', '') else ''
        }
        for container in product.get('containers', [])
        for price in container.get('prices', [])
    ]
    supplier_prices = product.get('supplierPrices', [])
    product_codes = [
        {
            "product_code": code.get('productCode', ''),
            "container_size": code.get('containerSize', '')
        }
        for code in product.get('codes', [])
    ]
    product_tags = product.get('tags', []) 

    price_iter = iter(product_prices)
    container_iter = iter(container_prices)
    supplier_iter = iter(supplier_prices)
    code_iter = iter(product_codes)
    tag_iter = iter(product_tags)

    rows_added = False

    while True:
        try:
            product_price = next(price_iter)
        except StopIteration:
            product_price = None

        try:
            container_price = next(container_iter)
        except StopIteration:
            container_price = None

        try:
            supplier_price = next(supplier_iter)
        except StopIteration:
            supplier_price = None

        try:
            product_code_entry = next(code_iter)
        except StopIteration:
            product_code_entry = None

        try:
            product_tag = next(tag_iter)
        except StopIteration:
            product_tag = None

        # Exit condition
        if not product_price and not container_price and not supplier_price and not product_code_entry and not product_tag:
            break

        # Append the row with detailed information
        row = [
            product_number,
            product_name,
            product_commodity_number,
            product_commodity_name,
            product_price.get('priceGroup', {}).get('number', '') if product_price else '',
            product_price.get('priceGroup', {}).get('name', '') if product_price else '',
            product_price.get('value', '') if product_price and product_price.get('priceGroup') else '',  # Price Group Value
            datetime.fromisoformat(product_price.get('validFrom', '')).strftime('%m/%d/%Y') if product_price and product_price.get('validFrom', '') else '',
            product_price.get('organizationalUnit', {}).get('number', '') if product_price else '',
            product_price.get('organizationalUnit', {}).get('name', '') if product_price else '',
            product_price.get('value', '') if product_price and product_price.get('organizationalUnit') else '',  # Organizational Unit Value
            container_price["container_number"] if container_price else '',
            container_price["container_name"] if container_price else '',
            container_price["container_price_value"] if container_price else '',
            container_price["container_price_group_number"] if container_price else '',
            container_price["default_container"] if container_price else '',
            container_price["container_valid"] if container_price else '',
            product_container_capacity,
            supplier_price.get('supplier', {}).get('number', '') if supplier_price else '',
            supplier_price.get('supplier', {}).get('name', '') if supplier_price else '',
            supplier_price.get('orderCode', '') if supplier_price else '',
            supplier_price.get('value', '') if supplier_price else '',
            supplier_price.get('containerSize', '') if supplier_price else '',
            product_code_entry["product_code"] if product_code_entry else '',
            product_code_entry["container_size"] if product_code_entry else '',
            product_tag.get('name', '') if product_tag else '',
            product_assortment_number,
            product_assortment_name,
            product_sector_number,
            product_sector_name,
            product_altSector_number,
            product_altSector_name,
            product_itemSequence_number,
            product_itemSequence_name,
            product_active,
            product_discountable,
            product_track_inventory
        ]

        sheet.append(row)

        # Apply fill to all columns
        for col in range(1, len(row) + 1):
            cell = sheet.cell(row=row_index, column=col)
            cell.fill = current_fill
            cell.alignment = Alignment(horizontal="left")
        row_index += 1
        rows_added = True

    # If no rows were added for prices, containers, suppliers, or codes, add a placeholder row
    if not rows_added:
        row = [
            product_number,
            product_name,
            product_commodity_number,
            product_commodity_name,
            '', '', '', '', '', '', '', '', '', '', '', '', '',
            product_container_capacity,
            '', '', '', '', '', '', '',
            product_active,
            product_discountable,
            product_track_inventory
        ]
        sheet.append(row)
        for col in range(1, len(row) + 1):
            cell = sheet.cell(row=row_index, column=col)
            cell.fill = current_fill
            cell.alignment = Alignment(horizontal="left")
        row_index += 1

    return row_index

# Main loop to fetch and process data
while True:
    json_data = fetch_data_from_api(koronaAccountId, page)
    if not json_data or not json_data.get('results', []):
        print(f"No data found for page {page}.")
        break

    for product in json_data.get('results', []):
        row_index = process_product(product, row_index, current_fill)
        current_fill = light_grey_fill if current_fill == white_fill else white_fill

    page += 1

# Adjust column widths and save the workbook
for col in sheet.columns:
    max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
    adjusted_width = max_length + 2
    sheet.column_dimensions[col[0].column_letter].width = adjusted_width

sheet.freeze_panes = sheet["A2"]
workbook.save('products.xlsx')
print("Data has been written to 'products.xlsx'")
